import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import logging
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import os
import json
import glob
import os
from datetime import datetime
from openpyxl import load_workbook

from ..crud import get_job_detail_metrics


# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
CVR_TEMPLATES_DIR = "CVR_Templates"
PROCESSED_CVR_DIR = "CVR_Processed"


class CVRUpdater:
    """
    CVR (Cost Value Reconciliation) updater
    Updates CVR Excel files with P&L and Invoice data
    """
    
    def __init__(self, cvr_template_path: str = "CVR_Templates/master_cvr.xlsx"):
        self.cvr_template_path = cvr_template_path
        self.update_rules = {}
        self.mapping_config = {}
        
    def load_update_rules(self, rules_file: str = "CVR_Templates/update_rules.json"):
        """Load update rules from JSON file"""
        try:
            if os.path.exists(rules_file):
                with open(rules_file, 'r') as f:
                    self.update_rules = json.load(f)
                logger.info(f"📋 Loaded update rules from {rules_file}")
            else:
                # Create default rules
                self.create_default_rules(rules_file)
        except Exception as e:
            logger.error(f"❌ Error loading update rules: {str(e)}")
            self.create_default_rules(rules_file)
    
    def create_default_rules(self, rules_file: str):
        """Create default update rules"""
        default_rules = {
            "cell_mappings": {
                "total_contract_value": "C3",
                "total_invoiced": "C4",
                "total_costs": "C5",
                "material_costs": "C6",
                "labour_costs": "C7",
                "plant_costs": "C8",
                "subcontract_costs": "C9",
                "projected_margin": "C10",
                "variations_approved": "C11",
                "variations_pending": "C12"
            },
            "cost_categories": {
                "material": ["Material", "Materials", "MATERIALS", "MAT"],
                "labour": ["Labour", "Labor", "LABOUR", "LAB", "Wages"],
                "plant": ["Plant", "PLANT", "Machinery", "Equipment", "EQUIP"],
                "subcontract": ["Subcontract", "SUB", "Subcontractor", "SC"]
            },
            "revenue_categories": {
                "main_contract": ["Main Contract", "Contract", "PRIMARY"],
                "variations": ["Variation", "VAR", "Additional Work", "EXTRA"]
            },
            "update_conditions": {
                "only_positive_values": True,
                "preserve_formulas": True,
                "backup_before_update": True,
                "log_all_changes": True
            }
        }
        
        # Save default rules
        os.makedirs(os.path.dirname(rules_file), exist_ok=True)
        with open(rules_file, 'w') as f:
            json.dump(default_rules, f, indent=2)
        
        self.update_rules = default_rules
        logger.info(f"📋 Created default update rules: {rules_file}")
    
    def update_cvr_with_pnl(self, cvr_file_path: str, pnl_data: Dict, job_code: str) -> Dict:
        """
        Update CVR file with P&L data
        
        Args:
            cvr_file_path: Path to CVR Excel file
            pnl_data: Parsed P&L data
            job_code: Job code to update
            
        Returns:
            Dictionary with update results
        """
        try:
            logger.info(f"📊 Updating CVR for job {job_code} with P&L data")
            
            # Load workbook
            wb = openpyxl.load_workbook(cvr_file_path)
            
            # Find or create job sheet
            sheet = self._get_or_create_job_sheet(wb, job_code)
            
            # Get job data from P&L
            if job_code not in pnl_data.get('data', {}):
                logger.warning(f"⚠️  Job {job_code} not found in P&L data")
                return {'success': False, 'error': f'Job {job_code} not found in P&L data'}
            
            job_data = pnl_data['data'][job_code]
            
            # Update cost cells
            updates = []
            
            # Total costs
            total_costs = job_data.get('total_expenses', 0)
            if self._update_cell(sheet, 'total_costs', total_costs):
                updates.append(f"Total costs: £{total_costs:,.2f}")
            
            # Category-wise costs
            expense_breakdown = job_data.get('expense_breakdown', {})
            
            # Material costs
            material_cost = self._calculate_category_cost(expense_breakdown, 'material')
            if self._update_cell(sheet, 'material_costs', material_cost):
                updates.append(f"Material costs: £{material_cost:,.2f}")
            
            # Labour costs
            labour_cost = self._calculate_category_cost(expense_breakdown, 'labour')
            if self._update_cell(sheet, 'labour_costs', labour_cost):
                updates.append(f"Labour costs: £{labour_cost:,.2f}")
            
            # Plant costs
            plant_cost = self._calculate_category_cost(expense_breakdown, 'plant')
            if self._update_cell(sheet, 'plant_costs', plant_cost):
                updates.append(f"Plant costs: £{plant_cost:,.2f}")
            
            # Subcontract costs
            subcontract_cost = self._calculate_category_cost(expense_breakdown, 'subcontract')
            if self._update_cell(sheet, 'subcontract_costs', subcontract_cost):
                updates.append(f"Subcontract costs: £{subcontract_cost:,.2f}")
            
            # Add timestamp
            sheet['A1'] = f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Save workbook
            wb.save(cvr_file_path)
            
            logger.info(f"✅ CVR updated successfully for job {job_code}")
            
            return {
                'success': True,
                'updates': updates,
                'job_code': job_code,
                'updated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"❌ Error updating CVR with P&L data: {str(e)}")
            return {'success': False, 'error': str(e)}
    
    def update_cvr_with_invoice(self, cvr_file_path: str, invoice_data: Dict, job_code: str) -> Dict:
        """
        Update CVR file with Invoice data
        
        Args:
            cvr_file_path: Path to CVR Excel file
            invoice_data: Parsed invoice data
            job_code: Job code to update
            
        Returns:
            Dictionary with update results
        """
        try:
            logger.info(f"📊 Updating CVR for job {job_code} with Invoice data")
            
            # Load workbook
            wb = openpyxl.load_workbook(cvr_file_path)
            
            # Find or create job sheet
            sheet = self._get_or_create_job_sheet(wb, job_code)
            
            # Get job data from invoices
            if job_code not in invoice_data.get('data', {}):
                logger.warning(f"⚠️  Job {job_code} not found in Invoice data")
                return {'success': False, 'error': f'Job {job_code} not found in Invoice data'}
            
            job_data = invoice_data['data'][job_code]
            
            # Update revenue cells
            updates = []
            
            # Total invoiced
            total_invoiced = job_data.get('total_invoiced', 0)
            if self._update_cell(sheet, 'total_invoiced', total_invoiced):
                updates.append(f"Total invoiced: £{total_invoiced:,.2f}")
            
            # Calculate projected margin (if contract value exists)
            contract_value_cell = self.update_rules.get('cell_mappings', {}).get('total_contract_value', 'C3')
            contract_value = sheet[contract_value_cell].value or 0
            
            if isinstance(contract_value, (int, float)) and contract_value > 0:
                # Get total costs from the sheet
                total_costs_cell = self.update_rules.get('cell_mappings', {}).get('total_costs', 'C5')
                total_costs = sheet[total_costs_cell].value or 0
                
                projected_margin = contract_value - total_costs
                if self._update_cell(sheet, 'projected_margin', projected_margin):
                    updates.append(f"Projected margin: £{projected_margin:,.2f}")
            
            # Add payment tracking info in additional columns
            sheet['E1'] = "Payment Tracking"
            sheet['E2'] = "Total Paid"
            sheet['F2'] = job_data.get('total_paid', 0)
            sheet['E3'] = "Outstanding Balance"
            sheet['F3'] = job_data.get('outstanding_balance', 0)
            sheet['E4'] = "Payment Rate"
            sheet['F4'] = f"{job_data.get('payment_rate', 0):.1f}%"
            
            # Add timestamp
            sheet['A1'] = f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Save workbook
            wb.save(cvr_file_path)
            
            logger.info(f"✅ CVR updated successfully for job {job_code}")
            
            return {
                'success': True,
                'updates': updates,
                'job_code': job_code,
                'updated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"❌ Error updating CVR with Invoice data: {str(e)}")
            return {'success': False, 'error': str(e)}
    
    def _get_or_create_job_sheet(self, wb: openpyxl.Workbook, job_code: str) -> openpyxl.worksheet.worksheet.Worksheet:
        """Get existing job sheet or create new one"""
        
        sheet_name = f"Job_{job_code}"
        
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        
        # Create new sheet from template
        if 'Template' in wb.sheetnames:
            template_sheet = wb['Template']
            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = sheet_name
        else:
            new_sheet = wb.create_sheet(sheet_name)
            self._setup_default_cvr_structure(new_sheet)
        
        return new_sheet
    
    def _setup_default_cvr_structure(self, sheet):
        """Setup default CVR structure in a new sheet"""
        
        # Headers
        sheet['A1'] = "Cost Value Reconciliation"
        sheet['A2'] = "Item"
        sheet['B2'] = "Description"
        sheet['C2'] = "Amount (£)"
        
        # Structure
        items = [
            ("Contract Value", "Original contract value"),
            ("Total Invoiced", "Total amount invoiced"),
            ("Total Costs", "Total costs incurred"),
            ("Material Costs", "Material and supplies"),
            ("Labour Costs", "Labour and wages"),
            ("Plant Costs", "Plant and equipment"),
            ("Subcontract Costs", "Subcontractor costs"),
            ("Projected Margin", "Projected profit margin"),
            ("Variations Approved", "Approved variations"),
            ("Variations Pending", "Pending variations")
        ]
        
        for i, (item, desc) in enumerate(items, start=3):
            sheet[f'A{i}'] = item
            sheet[f'B{i}'] = desc
            sheet[f'C{i}'] = 0
    
    def _update_cell(self, sheet, cell_key: str, value: float) -> bool:
        """Update a specific cell if conditions are met"""
        
        cell_address = self.update_rules.get('cell_mappings', {}).get(cell_key)
        if not cell_address:
            return False
        
        # Check update conditions
        conditions = self.update_rules.get('update_conditions', {})
        
        # Only update with positive values if required
        if conditions.get('only_positive_values', True) and value < 0:
            return False
        
        # Preserve formulas if required
        if conditions.get('preserve_formulas', True):
            current_value = sheet[cell_address].value
            if isinstance(current_value, str) and current_value.startswith('='):
                return False
        
        # Update the cell
        sheet[cell_address] = value
        return True
    
    def _calculate_category_cost(self, expense_breakdown: Dict, category: str) -> float:
        """Calculate cost for a specific category"""
        
        category_keywords = self.update_rules.get('cost_categories', {}).get(category, [])
        total_cost = 0
        
        for expense_type, amount in expense_breakdown.items():
            if any(keyword.lower() in expense_type.lower() for keyword in category_keywords):
                total_cost += amount
        
        return total_cost
    
    def update_multiple_jobs(self, cvr_file_path: str, pnl_data: Dict, invoice_data: Dict) -> Dict:
        """
        Update CVR for multiple jobs
        
        Args:
            cvr_file_path: Path to CVR Excel file
            pnl_data: Parsed P&L data
            invoice_data: Parsed invoice data
            
        Returns:
            Dictionary with update results for all jobs
        """
        results = {
            'success': True,
            'job_results': {},
            'errors': [],
            'updated_jobs': [],
            'updated_at': datetime.now().isoformat()
        }
        
        # Get all unique job codes
        pnl_jobs = set(pnl_data.get('data', {}).keys())
        invoice_jobs = set(invoice_data.get('data', {}).keys())
        all_jobs = pnl_jobs.union(invoice_jobs)
        
        for job_code in all_jobs:
            job_result = {'job_code': job_code, 'pnl_updated': False, 'invoice_updated': False}
            
            # Update with P&L data
            if job_code in pnl_jobs:
                pnl_result = self.update_cvr_with_pnl(cvr_file_path, pnl_data, job_code)
                job_result['pnl_updated'] = pnl_result['success']
                if not pnl_result['success']:
                    results['errors'].append(f"P&L update failed for {job_code}: {pnl_result.get('error', '')}")
            
            # Update with Invoice data
            if job_code in invoice_jobs:
                invoice_result = self.update_cvr_with_invoice(cvr_file_path, invoice_data, job_code)
                job_result['invoice_updated'] = invoice_result['success']
                if not invoice_result['success']:
                    results['errors'].append(f"Invoice update failed for {job_code}: {invoice_result.get('error', '')}")
            
            results['job_results'][job_code] = job_result
            
            if job_result['pnl_updated'] or job_result['invoice_updated']:
                results['updated_jobs'].append(job_code)
        
        # Overall success if no errors
        results['success'] = len(results['errors']) == 0
        
        logger.info(f"✅ CVR batch update completed. Updated {len(results['updated_jobs'])} jobs")
        
        return results
    
    def get_cvr_dashboard_data(self, cvr_file_path: str, job_code: str) -> Dict:
        """
        Extract dashboard data from CVR file
        
        Args:
            cvr_file_path: Path to CVR Excel file
            job_code: Job code to extract data for
            
        Returns:
            Dictionary with dashboard data
        """
        try:
            wb = openpyxl.load_workbook(cvr_file_path)
            sheet_name = f"Job_{job_code}"
            
            if sheet_name not in wb.sheetnames:
                return {'success': False, 'error': f'Sheet for job {job_code} not found'}
            
            sheet = wb[sheet_name]
            
            # Extract data from cells
            data = {}
            for key, cell_address in self.update_rules.get('cell_mappings', {}).items():
                try:
                    value = sheet[cell_address].value
                    if value is None:
                        value = 0
                    elif isinstance(value, str) and value.startswith('='):
                        # Handle formula cells
                        value = 0
                    data[key] = float(value) if isinstance(value, (int, float)) else 0
                except Exception as e:
                    logger.warning(f"⚠️  Error reading cell {cell_address}: {str(e)}")
                    data[key] = 0
            
            # Calculate additional metrics
            contract_value = data.get('total_contract_value', 0)
            total_costs = data.get('total_costs', 0)
            total_invoiced = data.get('total_invoiced', 0)
            
            # Calculate margins and percentages
            data['cost_percentage'] = (total_costs / contract_value * 100) if contract_value > 0 else 0
            data['invoiced_percentage'] = (total_invoiced / contract_value * 100) if contract_value > 0 else 0
            data['gross_margin'] = total_invoiced - total_costs
            data['gross_margin_percentage'] = (data['gross_margin'] / total_invoiced * 100) if total_invoiced > 0 else 0
            
            # Payment tracking data
            try:
                data['total_paid'] = sheet['F2'].value or 0
                data['outstanding_balance'] = sheet['F3'].value or 0
                payment_rate_cell = sheet['F4'].value
                if isinstance(payment_rate_cell, str) and '%' in payment_rate_cell:
                    data['payment_rate'] = float(payment_rate_cell.replace('%', ''))
                else:
                    data['payment_rate'] = float(payment_rate_cell) if payment_rate_cell else 0
            except:
                data['total_paid'] = 0
                data['outstanding_balance'] = total_invoiced
                data['payment_rate'] = 0
            
            # Last updated timestamp
            try:
                last_updated = sheet['A1'].value
                if isinstance(last_updated, str) and 'Last updated:' in last_updated:
                    data['last_updated'] = last_updated.replace('Last updated: ', '')
                else:
                    data['last_updated'] = 'Unknown'
            except:
                data['last_updated'] = 'Unknown'
            
            return {
                'success': True,
                'job_code': job_code,
                'data': data
            }
            
        except Exception as e:
            logger.error(f"❌ Error extracting CVR dashboard data: {str(e)}")
            return {'success': False, 'error': str(e)}
    
    def get_all_jobs_summary(self, cvr_file_path: str) -> Dict:
        """
        Get summary data for all jobs in CVR file
        
        Args:
            cvr_file_path: Path to CVR Excel file
            
        Returns:
            Dictionary with summary data for all jobs
        """
        try:
            wb = openpyxl.load_workbook(cvr_file_path)
            
            # Find all job sheets
            job_sheets = [sheet for sheet in wb.sheetnames if sheet.startswith('Job_')]
            
            summary_data = {
                'total_jobs': len(job_sheets),
                'total_contract_value': 0,
                'total_invoiced': 0,
                'total_costs': 0,
                'total_margin': 0,
                'jobs': []
            }
            
            for sheet_name in job_sheets:
                job_code = sheet_name.replace('Job_', '')
                job_data = self.get_cvr_dashboard_data(cvr_file_path, job_code)
                
                if job_data['success']:
                    data = job_data['data']
                    
                    # Add to totals
                    summary_data['total_contract_value'] += data.get('total_contract_value', 0)
                    summary_data['total_invoiced'] += data.get('total_invoiced', 0)
                    summary_data['total_costs'] += data.get('total_costs', 0)
                    summary_data['total_margin'] += data.get('gross_margin', 0)
                    
                    # Add job summary
                    summary_data['jobs'].append({
                        'job_code': job_code,
                        'contract_value': data.get('total_contract_value', 0),
                        'invoiced': data.get('total_invoiced', 0),
                        'costs': data.get('total_costs', 0),
                        'margin': data.get('gross_margin', 0),
                        'margin_percentage': data.get('gross_margin_percentage', 0),
                        'last_updated': data.get('last_updated', 'Unknown')
                    })
            
            # Calculate overall percentages
            if summary_data['total_contract_value'] > 0:
                summary_data['overall_cost_percentage'] = (summary_data['total_costs'] / summary_data['total_contract_value']) * 100
                summary_data['overall_invoiced_percentage'] = (summary_data['total_invoiced'] / summary_data['total_contract_value']) * 100
            else:
                summary_data['overall_cost_percentage'] = 0
                summary_data['overall_invoiced_percentage'] = 0
            
            if summary_data['total_invoiced'] > 0:
                summary_data['overall_margin_percentage'] = (summary_data['total_margin'] / summary_data['total_invoiced']) * 100
            else:
                summary_data['overall_margin_percentage'] = 0
            
            return {
                'success': True,
                'summary': summary_data
            }
            
        except Exception as e:
            logger.error(f"❌ Error getting jobs summary: {str(e)}")
            return {'success': False, 'error': str(e)}
    
    def create_master_cvr_template(self, template_path: str = "CVR_Templates/master_cvr.xlsx"):
        """Create a master CVR template file"""
        try:
            os.makedirs(os.path.dirname(template_path), exist_ok=True)
            
            wb = openpyxl.Workbook()
            
            # Create Template sheet
            template_sheet = wb.active
            template_sheet.title = "Template"
            self._setup_default_cvr_structure(template_sheet)
            
            # Create Summary sheet
            summary_sheet = wb.create_sheet("Summary")
            summary_sheet['A1'] = "NDA Company - CVR Summary"
            summary_sheet['A2'] = "Generated on:"
            summary_sheet['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Summary headers
            summary_sheet['A4'] = "Job Code"
            summary_sheet['B4'] = "Contract Value"
            summary_sheet['C4'] = "Total Invoiced"
            summary_sheet['D4'] = "Total Costs"
            summary_sheet['E4'] = "Margin"
            summary_sheet['F4'] = "Margin %"
            summary_sheet['G4'] = "Status"
            
            # Save template
            wb.save(template_path)
            
            logger.info(f"✅ Master CVR template created: {template_path}")
            return {'success': True, 'template_path': template_path}
            
        except Exception as e:
            logger.error(f"❌ Error creating CVR template: {str(e)}")
            return {'success': False, 'error': str(e)}
        
    def validate_cvr_structure(filepath: str) -> dict:

        wb = load_workbook(filepath, read_only=True)
        sheet = wb.active
        headers = [cell.value for cell in next(sheet.rows)]
        required = ["Job Code", "Cost to Date", "Invoiced", "Est. Final Cost", "Amended Value", "Margin"]
        missing = [h for h in required if h not in headers]
        return {"valid": not missing, "missing_headers": missing}
    
    def process_all_jobs_cvr(db, template_path: str = None) -> dict:

        """
        1. Load the master template (latest if not passed)
        2. For each data row, look up metrics from DB
        3. Overwrite columns: Cost to Date, Invoiced, Est. Final Cost, Amended Value, Margin
        4. Save new file under CVR_Processed with timestamp
        """
        os.makedirs(PROCESSED_CVR_DIR, exist_ok=True)
        # pick latest if no explicit
        if not template_path:
            files = sorted(glob.glob(f"{CVR_TEMPLATES_DIR}/*.xlsx"))
            if not files:
                raise FileNotFoundError("No CVR template found")
            template_path = files[-1]

        wb = load_workbook(template_path)
        sheet = wb.active
        # find column indices by header row
        col_index = {cell.value: idx + 1 for idx, cell in enumerate(sheet[1])}

        for row in sheet.iter_rows(min_row=2):
            job_code = row[col_index["Job Code"] - 1].value
            # fetch metrics
            metrics = get_job_detail_metrics(db, job_code=job_code)  # adjust CRUD signature
            if not metrics:
                continue

            # write values
            row[col_index["Cost to Date"] - 1].value = metrics["total_costs"]
            row[col_index["Invoiced"] - 1].value = metrics["total_invoiced"]
            row[col_index["Est. Final Cost"] - 1].value = metrics["total_costs"] + metrics["pending_invoices"]
            row[col_index["Amended Value"] - 1].value = metrics["amended_value"]
            row[col_index["Margin"] - 1].value = metrics["projected_margin"]

        out_name = f"cvr_processed_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        out_path = os.path.join(PROCESSED_CVR_DIR, out_name)
        wb.save(out_path)
        return {"file": out_name, "path": out_path}
    
    def download_latest_cvr() -> str:
        files = sorted(glob.glob(f"{PROCESSED_CVR_DIR}/*.xlsx"))
        if not files:
            raise FileNotFoundError("No processed CVR found")
        return files[-1]


